import openpyxl, unicodedata, re, json, os, sys
sys.stdout.reconfigure(encoding='utf-8')
SKIP=['resumen']
def slugify(t):
    t=t.lower().strip();t=unicodedata.normalize('NFD',t);t=''.join(c for c in t if unicodedata.category(c)!='Mn');t=re.sub(r'[^a-z0-9]+','-',t);return t.strip('-')
def parse_bool(v):
    if isinstance(v,bool):return v
    if v is None:return False
    return str(v).lower() in['si','yes','true','✅','🔥']
def parse_int(v):
    if v is None:return 0
    try:return int(re.sub(r'[^0-9]','',str(v))or'0')
    except:return 0
wb=openpyxl.load_workbook('data/tgonly-grupos.xlsx')
categories=[];groups=[]
for sheet_name in wb.sheetnames:
    if sheet_name.lower() in SKIP:continue
    ws=wb[sheet_name];rows=list(ws.iter_rows(values_only=True))
    header_row=-1
    for i,row in enumerate(rows[:5]):
        if row and any(str(c or'').lower() in['nombre','nombre del grupo'] for c in row):header_row=i;break
    if header_row==-1:continue
    category_slug=slugify(sheet_name);data_rows=[r for r in rows[header_row+1:] if r and r[2]]
    if not data_rows:continue
    valid=[r for r in data_rows if str(r[2]or'').strip()]
    categories.append({'emoji':str(data_rows[0][1]or'📁').strip(),'name':sheet_name,'count':str(len(valid)),'slug':category_slug})
    for row in data_rows:
        name=str(row[2]or'').strip()
        if not name:continue
        tags=[t.strip() for t in str(row[6]or'').split('|') if t.strip()]
        link=str(row[9]or'').strip()
        groups.append({'emoji':str(row[1]or'').strip(),'color':'blue','name':name,'members':str(row[3]or'0'),'verified':parse_bool(row[4]),'desc':str(row[5]or'').strip(),'tags':tags,'trending':parse_bool(row[7]),'category':category_slug,'link':link if link and link!='#' else '#','score':parse_int(row[8]) or 50})
    print(f"  {sheet_name} -> /{category_slug} ({len(valid)} grupos)")
lines=["// AUTO-GENERADO desde data/tgonly-grupos.xlsx",""]
lines+=["export type Category = { emoji: string; name: string; count: string; slug: string }",""]
lines+=["export type Group = { emoji: string; color: string; name: string; members: string; verified: boolean; desc: string; tags: string[]; trending: boolean; category: string; link: string; score?: number }",""]
lines.append("export const categories: Category[] = [")
for cat in categories:lines.append(f"  {{ emoji: {json.dumps(cat['emoji'])}, name: {json.dumps(cat['name'])}, count: {json.dumps(cat['count'])}, slug: {json.dumps(cat['slug'])} }},")
lines.append("]")
lines.append("")
lines.append("export const groups: Group[] = [")
for g in groups:lines.append(f"  {{ emoji: {json.dumps(g['emoji'])}, color: 'blue', name: {json.dumps(g['name'])}, members: {json.dumps(str(g['members']))}, verified: {'true' if g['verified'] else 'false'}, desc: {json.dumps(g['desc'])}, tags: {json.dumps(g['tags'])}, trending: {'true' if g['trending'] else 'false'}, category: {json.dumps(g['category'])}, link: {json.dumps(g['link'])}, score: {g['score']} }},")
lines.append("]")
with open('data/groups.ts','w',encoding='utf-8') as f:f.write('\n'.join(lines))
print(f"\nGenerado: {len(categories)} categorias, {len(groups)} grupos -> data/groups.ts")
