import openpyxl
import requests
import json
import unicodedata
import re
import os
import sys

SUPABASE_URL = os.environ['SUPABASE_URL']
SERVICE_KEY  = os.environ['SUPABASE_SERVICE_KEY']

HEADERS = {
    "apikey": SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type": "application/json",
    "Prefer": "resolution=merge-duplicates"
}

SKIP_SHEETS = ['resumen']

def slugify(text):
    text = text.lower().strip()
    text = unicodedata.normalize('NFD', text)
    text = ''.join(c for c in text if unicodedata.category(c) != 'Mn')
    text = re.sub(r'[^a-z0-9]+', '-', text)
    return text.strip('-')

def parse_bool(val):
    if isinstance(val, bool): return val
    if val is None: return False
    s = str(val).lower()
    return any(x in s for x in ['si', 'yes', 'true', '✅', '🔥'])

def parse_int(val):
    if val is None: return 0
    return int(re.sub(r'[^0-9]', '', str(val)) or '0')

# Verificar conexion
print("Conectando a Supabase...")
r = requests.get(f"{SUPABASE_URL}/rest/v1/groups?limit=1", headers=HEADERS)
if r.status_code not in [200, 206]:
    print(f"ERROR: {r.status_code} — {r.text[:300]}")
    sys.exit(1)
print("Conexion OK")

# Leer Excel
excel_path = 'data/tgonly-grupos.xlsx'
wb = openpyxl.load_workbook(excel_path)
print(f"Hojas encontradas: {wb.sheetnames}")

total_ok = 0
total_err = 0

for sheet_name in wb.sheetnames:
    if sheet_name.lower() in SKIP_SHEETS:
        print(f"Saltando: {sheet_name}")
        continue

    ws = wb[sheet_name]
    rows = list(ws.iter_rows(values_only=True))

    # Encontrar fila de headers
    header_row = -1
    for i, row in enumerate(rows[:5]):
        if row and any(str(c or '').lower() in ['nombre', 'nombre del grupo'] for c in row):
            header_row = i
            break

    if header_row == -1:
        print(f"'{sheet_name}' — sin headers reconocibles, saltando")
        continue

    category_slug = slugify(sheet_name)
    data_rows = [r for r in rows[header_row+1:] if r and r[2]]

    if not data_rows:
        print(f"'{sheet_name}' — sin datos")
        continue

    print(f"{sheet_name} -> /{category_slug} ({len(data_rows)} grupos)")

    groups = []
    for row in data_rows:
        name = str(row[2] or '').strip()
        if not name:
            continue
        tags = [t.strip() for t in str(row[6] or '').split('|') if t.strip()]
        groups.append({
            "name":     name,
            "emoji":    str(row[1] or '').strip(),
            "members":  parse_int(row[3]),
            "verified": parse_bool(row[4]),
            "desc":     str(row[5] or '').strip(),
            "tags":     tags,
            "trending": parse_bool(row[7]),
            "score":    parse_int(row[8]),
            "link":     str(row[9] or '').strip(),
            "category": category_slug,
            "color":    "blue",
        })

    if not groups:
        continue

    for i in range(0, len(groups), 50):
        batch = groups[i:i+50]
        r = requests.post(
            f"{SUPABASE_URL}/rest/v1/groups",
            headers=HEADERS,
            data=json.dumps(batch)
        )
        if r.status_code in [200, 201]:
            print(f"  OK: {len(batch)} grupos subidos")
            total_ok += len(batch)
        else:
            print(f"  ERROR {r.status_code}: {r.text[:300]}")
            total_err += 1

print(f"\nTotal: {total_ok} grupos importados, {total_err} errores")
if total_err > 0:
    sys.exit(1)
